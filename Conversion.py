from datetime import datetime
from datetime import timedelta
from pysb import SbApi
import openpyxl as xl
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import shutil

# token = 'eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJqdGkiOiI5NmYzZTEwYy03MmUxLTRlYTUtYmU4NS1mMjc3YWZlYWNkNjEiLCJpYXQiOjE1NTA4MDUxNjYsInN1YiI6MTAwMDE2LCJhdWQiOjc5MTYsImlzcyI6bnVsbH0.4ctJjZuAjQpTzabEdAnUcVFivZdKPLTfESJSpHZz1_Y'
# sb_url = 'rsr'

workbook = load_workbook(filename= "purchase-order.xlsx", data_only=True)

sheet = workbook.active

# PO Data
po_num = sheet["G3"].value
po_vend = sheet["A9"].value
po_start_ship = sheet["G2"].value
po_end_ship = po_start_ship + timedelta(days=60)

# Item Data 
Item_number = []
Description = []
Item_QTY = []
Unit_Cost = []

# Item Numbers
for row in sheet.iter_rows(min_row=20, min_col=1, max_col=1):
    if row[0].value is None:
        break
    Item_number.append(row[0].value)

# Item Description
for row in sheet.iter_rows(min_row=20, min_col=2, max_col=2):
    if row[0].value is None:
        break
    Description.append(row[0].value)

# Item QTY
for row in sheet.iter_rows(min_row=20, min_col=5, max_col=5):
    if row[0].value is None:
        break
    Item_QTY.append(row[0].value)

# Item Unit Price
for row in sheet.iter_rows(min_row=20, min_col=6, max_col=6):
    if row[0].value is None:
        break
    Unit_Cost.append(row[0].value)

File_name = input("Whats the New File Name?")
New_File = shutil.copy("PO Import Template.xlsx", File_name + ".xlsx") 

New_wb = xl.load_workbook(New_File)
New_ws = New_wb.active

New_ws["A2"] = po_num
New_ws["C2"] = po_start_ship
New_ws["D2"] = po_end_ship

for I in Item_number:
    New_ws.cell(row = 2 + I, column = 9).value = Item_number[I]

#test
    
New_wb.save(str(New_File))

New_ws["j2"] = Description
New_ws["H2"] = Item_QTY
New_ws["G2"] = Unit_Cost



print(Item_number)
print(Description)
print(Item_QTY)
print(Unit_Cost)
print(po_num)
print(po_start_ship)
print(po_end_ship)
